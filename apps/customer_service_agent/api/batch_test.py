"""批量测试 API。

支持从 Excel 导入问题 → 后台批量调用 QA → 人工审核准确率。

Excel 格式（.xlsx，第一行为表头，列名灵活匹配）：
| 问题 | 期望答案 | 分类 | 备注 |
|------|----------|------|------|

API 列表：
- POST   /admin/api/batch-test/upload    上传 Excel，解析问题
- POST   /admin/api/batch-test/{id}/run   启动批量测试（后台异步）
- GET    /admin/api/batch-test             任务列表
- GET    /admin/api/batch-test/{id}/status 任务进度
- GET    /admin/api/batch-test/{id}/results 结果列表（分页）
- PUT    /admin/api/batch-test/items/{id}  审核单条结果
- DELETE /admin/api/batch-test/{id}        删除任务
- GET    /admin/api/batch-test/template     下载 Excel 模板
"""

from __future__ import annotations

import asyncio
import io
import time
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select, update
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from apps.customer_service_agent.agent.qa import QAService
from common.config.config import settings
from common.database.database import BatchTestItem, BatchTestTask
from common.logger.logger import get_logger

router = APIRouter(prefix="/admin/api/batch-test", tags=["batch-test"])
log = get_logger(__name__)

# ===== 共享 DB 引擎（避免每个请求/后台任务创建/销毁连接池）=====

_shared_engine: Any = None
_shared_session_maker: Any = None


def _get_db():
    """获取共享的异步 DB session maker，惰性初始化。"""
    global _shared_engine, _shared_session_maker
    if _shared_engine is None:
        _shared_engine = create_async_engine(settings.database_url, echo=False)
        _shared_session_maker = async_sessionmaker(
            _shared_engine, expire_on_commit=False
        )
    return _shared_session_maker


# 列名映射：支持多种常见表头写法
COLUMN_ALIASES = {
    "question": ["问题", "提问", "用户问题", "question", "query", "Q"],
    "expected_answer": ["期望答案", "标准答案", "参考答案", "正确答案", "expected", "answer", "A"],
    "category": ["分类", "类别", "类型", "category", "tag"],
    "note": ["备注", "说明", "note", "remark"],
}


def _match_column(header: str) -> str | None:
    """将表头匹配到标准列名。"""
    h = header.strip().lower()
    for std, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            if h == a.lower():
                return std
    return None


def _parse_excel(data: bytes) -> list[dict[str, str]]:
    """解析 Excel，返回 [{question, expected_answer, category, note}, ...]。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        raise ValueError("Excel 文件为空")

    # 匹配表头
    header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_map: dict[int, str] = {}
    for i, h in enumerate(header_row):
        std = _match_column(h)
        if std:
            col_map[i] = std

    # 至少要有"问题"列
    if "question" not in col_map.values():
        wb.close()
        raise ValueError(
            f"未找到问题列，请确保表头包含: {COLUMN_ALIASES['question']}"
        )

    items = []
    for row in rows[1:]:
        if not row:
            continue
        item: dict[str, str] = {}
        for i, val in enumerate(row):
            if i not in col_map:
                continue
            std = col_map[i]
            if val is not None and str(val).strip():
                item[std] = str(val).strip()
        # 必须有问题内容
        if item.get("question"):
            items.append(item)

    wb.close()
    return items


# ===== Pydantic =====


class ReviewUpdate(BaseModel):
    review_status: str = Field(..., pattern="^(correct|incorrect|pending)$")
    review_reason: str | None = None


# ===== API =====


@router.get("/template")
async def download_template() -> StreamingResponse:
    """下载 Excel 模板。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "批量测试"
    ws.append(["问题", "期望答案", "分类", "备注"])
    # 示例数据
    ws.append(["发货时间是多久？", "48小时内发货，默认中通快递", "物流", ""])
    ws.append(["这个色号适合黄皮吗？", "#12号色适合黄皮，能提亮肤色", "产品咨询", ""])
    ws.append(["支持七天无理由退货吗？", "支持，不影响二次销售可退", "售后", ""])
    ws.append(["怎么修改收货地址？", "发货前联系客服修改", "订单", ""])

    # 设置列宽
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=batch_test_template.xlsx"},
    )


@router.post("/upload")
async def upload_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    """上传 Excel 文件，解析问题并创建测试任务。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 格式的 Excel 文件")

    content = await file.read()
    try:
        items = _parse_excel(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        log.error("batch_test.parse_failed", error=str(e))
        raise HTTPException(400, f"Excel 解析失败: {e}")

    if not items:
        raise HTTPException(400, "未解析到有效问题，请检查 Excel 内容")

    task_id = str(uuid4())
    db_maker = _get_db()

    async with db_maker() as session:
        # 创建任务
        session.add(BatchTestTask(
            id=task_id,
            filename=file.filename,
            total=len(items),
            status="pending",
        ))
        # 创建条目
        for i, item in enumerate(items):
            session.add(BatchTestItem(
                task_id=task_id,
                seq=i + 1,
                question=item["question"],
                expected_answer=item.get("expected_answer"),
                category=item.get("category"),
            ))
        await session.commit()

    log.info("batch_test.uploaded", task_id=task_id, total=len(items))
    return {"task_id": task_id, "total": len(items), "filename": file.filename}


@router.post("/{task_id}/run")
async def run_batch_test(task_id: str) -> dict[str, Any]:
    """启动批量测试（后台异步执行）。"""
    db_maker = _get_db()

    async with db_maker() as session:
        task = await session.get(BatchTestTask, task_id)
        if not task:
            raise HTTPException(404, "任务不存在")
        if task.status == "running":
            raise HTTPException(409, "任务正在运行中")
        if task.status == "completed":
            raise HTTPException(409, "任务已完成，请重新上传")
        await session.execute(
            update(BatchTestTask).where(BatchTestTask.id == task_id).values(status="running")
        )
        await session.commit()

    # 后台执行（添加异常回调，避免 "Task exception was never retrieved" 警告）
    bg_task = asyncio.create_task(_run_test_background(task_id))
    bg_task.add_done_callback(_bg_task_callback)

    return {"task_id": task_id, "status": "running"}


def _bg_task_callback(task: asyncio.Task) -> None:
    """后台任务完成回调，记录未捕获异常。"""
    if task.cancelled():
        return
    exc = task.exception()
    if exc:
        log.error("batch_test.bg_task_crashed", error=str(exc))


async def _run_test_background(task_id: str) -> None:
    """后台执行批量测试（使用共享 DB 引擎）。"""
    db_maker = _get_db()
    qa = QAService()
    completed = 0

    try:
        async with db_maker() as session:
            result = await session.execute(
                select(BatchTestItem)
                .where(BatchTestItem.task_id == task_id)
                .where(BatchTestItem.test_status == "pending")
                .order_by(BatchTestItem.seq)
            )
            items = result.scalars().all()

        for item in items:
            t0 = time.monotonic()
            try:
                resp = await qa.answer(item.question)
                elapsed_ms = int((time.monotonic() - t0) * 1000)
                sources_text = "\n".join(
                    f"[{h.collection}] {h.text[:100]}" for h in resp.sources
                ) if resp.sources else ""

                async with db_maker() as session:
                    await session.execute(
                        update(BatchTestItem).where(BatchTestItem.id == item.id).values(
                            actual_answer=resp.answer,
                            matched=resp.matched,
                            sources_count=len(resp.sources),
                            sources_text=sources_text,
                            response_time_ms=elapsed_ms,
                            test_status="done",
                            error_msg=None,
                        )
                    )
                    await session.execute(
                        update(BatchTestTask).where(BatchTestTask.id == task_id).values(
                            completed=completed + 1
                        )
                    )
                    await session.commit()

            except Exception as e:
                log.error("batch_test.item_failed", item_id=item.id, error=str(e))
                async with db_maker() as session:
                    await session.execute(
                        update(BatchTestItem).where(BatchTestItem.id == item.id).values(
                            test_status="error",
                            error_msg=str(e)[:500],
                        )
                    )
                    await session.execute(
                        update(BatchTestTask).where(BatchTestTask.id == task_id).values(
                            completed=completed + 1
                        )
                    )
                    await session.commit()

            completed += 1

        # 标记完成
        async with db_maker() as session:
            await session.execute(
                update(BatchTestTask).where(BatchTestTask.id == task_id).values(status="completed")
            )
            await session.commit()

        log.info("batch_test.completed", task_id=task_id, completed=completed)

    except Exception as e:
        log.error("batch_test.background_failed", task_id=task_id, error=str(e))
        async with db_maker() as session:
            await session.execute(
                update(BatchTestTask).where(BatchTestTask.id == task_id).values(status="failed")
            )
            await session.commit()


@router.get("")
async def list_tasks(
    limit: int = Query(default=20, le=100),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    """列出所有测试任务。"""
    db_maker = _get_db()
    async with db_maker() as session:
        # 总数
        count_r = await session.execute(select(func.count()).select_from(BatchTestTask))
        total = count_r.scalar() or 0

        # 分页列表
        result = await session.execute(
            select(BatchTestTask)
            .order_by(BatchTestTask.created_at.desc())
            .offset(offset)
            .limit(limit)
        )
        tasks = result.scalars().all()

        # 一次性获取所有任务的审核统计（避免 N+1）
        task_ids = [t.id for t in tasks]
        review_stats = {}
        if task_ids:
            stats_query = (
                select(
                    BatchTestItem.task_id,
                    BatchTestItem.review_status,
                    func.count().label("cnt"),
                )
                .where(BatchTestItem.task_id.in_(task_ids))
                .group_by(BatchTestItem.task_id, BatchTestItem.review_status)
            )
            stats_result = await session.execute(stats_query)
            for row in stats_result:
                if row.task_id not in review_stats:
                    review_stats[row.task_id] = {"correct": 0, "incorrect": 0}
                if row.review_status in ("correct", "incorrect"):
                    review_stats[row.task_id][row.review_status] = row.cnt

        items = []
        for t in tasks:
            stats = review_stats.get(t.id, {"correct": 0, "incorrect": 0})
            correct = stats["correct"]
            incorrect = stats["incorrect"]
            reviewed = correct + incorrect
            accuracy = round(correct / reviewed * 100, 1) if reviewed > 0 else None

            items.append({
                "id": t.id,
                "filename": t.filename,
                "total": t.total,
                "completed": t.completed,
                "status": t.status,
                "reviewed": reviewed,
                "correct": correct,
                "incorrect": incorrect,
                "accuracy": accuracy,
                "created_at": t.created_at.isoformat() if t.created_at else None,
            })

    return {"items": items, "total": total, "limit": limit, "offset": offset}


@router.get("/{task_id}/status")
async def get_status(task_id: str) -> dict[str, Any]:
    """获取任务进度。"""
    db_maker = _get_db()
    async with db_maker() as session:
        task = await session.get(BatchTestTask, task_id)
        if not task:
            raise HTTPException(404, "任务不存在")

        # 统计
        done_r = await session.execute(
            select(func.count()).select_from(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .where(BatchTestItem.test_status == "done")
        )
        done = done_r.scalar() or 0

        error_r = await session.execute(
            select(func.count()).select_from(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .where(BatchTestItem.test_status == "error")
        )
        errors = error_r.scalar() or 0

        matched_r = await session.execute(
            select(func.count()).select_from(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .where(BatchTestItem.matched == True)  # noqa: E712
        )
        matched = matched_r.scalar() or 0

        correct_r = await session.execute(
            select(func.count()).select_from(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .where(BatchTestItem.review_status == "correct")
        )
        correct = correct_r.scalar() or 0

        incorrect_r = await session.execute(
            select(func.count()).select_from(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .where(BatchTestItem.review_status == "incorrect")
        )
        incorrect = incorrect_r.scalar() or 0

        pending_review = task.completed - correct - incorrect

    return {
        "task_id": task_id,
        "status": task.status,
        "total": task.total,
        "completed": task.completed,
        "done": done,
        "errors": errors,
        "matched": matched,
        "unmatched": done - matched,
        "reviewed": correct + incorrect,
        "correct": correct,
        "incorrect": incorrect,
        "pending_review": pending_review,
    }


@router.get("/{task_id}/results")
async def get_results(
    task_id: str,
    review_filter: str = Query(default="", pattern="^(|correct|incorrect|pending)$"),
    limit: int = Query(default=50, le=200),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    """获取测试结果（分页，可按审核状态筛选）。"""
    db_maker = _get_db()
    async with db_maker() as session:
        task = await session.get(BatchTestTask, task_id)
        if not task:
            raise HTTPException(404, "任务不存在")

        # 构建查询
        base_query = select(BatchTestItem).where(BatchTestItem.task_id == task_id)
        count_query = select(func.count()).select_from(BatchTestItem).where(BatchTestItem.task_id == task_id)

        if review_filter:
            base_query = base_query.where(BatchTestItem.review_status == review_filter)
            count_query = count_query.where(BatchTestItem.review_status == review_filter)

        # 总数
        total_r = await session.execute(count_query)
        total = total_r.scalar() or 0

        # 分页
        result = await session.execute(
            base_query.order_by(BatchTestItem.seq).offset(offset).limit(limit)
        )
        items = result.scalars().all()

        data = []
        for it in items:
            data.append({
                "id": it.id,
                "seq": it.seq,
                "question": it.question,
                "expected_answer": it.expected_answer,
                "category": it.category,
                "actual_answer": it.actual_answer,
                "matched": it.matched,
                "sources_count": it.sources_count,
                "sources_text": it.sources_text,
                "response_time_ms": it.response_time_ms,
                "test_status": it.test_status,
                "error_msg": it.error_msg,
                "review_status": it.review_status,
                "review_reason": it.review_reason,
            })
    return {"items": data, "total": total, "limit": limit, "offset": offset}


@router.put("/items/{item_id}")
async def review_item(item_id: int, body: ReviewUpdate) -> dict[str, Any]:
    """审核单条结果：标记正确/错误 + 原因。"""
    db_maker = _get_db()
    async with db_maker() as session:
        item = await session.get(BatchTestItem, item_id)
        if not item:
            raise HTTPException(404, "条目不存在")

        item.review_status = body.review_status
        item.review_reason = body.review_reason
        await session.commit()

    log.info("batch_test.reviewed", item_id=item_id, status=body.review_status)
    return {"id": item_id, "review_status": body.review_status, "review_reason": body.review_reason}


@router.delete("/{task_id}")
async def delete_task(task_id: str) -> dict[str, Any]:
    """删除测试任务及其所有条目。"""
    db_maker = _get_db()
    async with db_maker() as session:
        task = await session.get(BatchTestTask, task_id)
        if not task:
            raise HTTPException(404, "任务不存在")

        await session.execute(
            delete(BatchTestItem).where(BatchTestItem.task_id == task_id)
        )
        await session.execute(
            delete(BatchTestTask).where(BatchTestTask.id == task_id)
        )
        await session.commit()

    log.info("batch_test.deleted", task_id=task_id)
    return {"deleted": task_id}


@router.get("/{task_id}/export")
async def export_results(task_id: str) -> StreamingResponse:
    """导出测试结果为 Excel（含审核列）。"""
    from openpyxl import Workbook

    db_maker = _get_db()
    async with db_maker() as session:
        task = await session.get(BatchTestTask, task_id)
        if not task:
            raise HTTPException(404, "任务不存在")

        result = await session.execute(
            select(BatchTestItem)
            .where(BatchTestItem.task_id == task_id)
            .order_by(BatchTestItem.seq)
        )
        items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "测试结果"
    headers = [
        "序号", "问题", "期望答案", "分类",
        "AI回答", "是否命中", "来源数", "耗时(ms)",
        "测试状态", "错误信息",
        "审核结果", "未答对原因",
    ]
    ws.append(headers)

    for it in items:
        ws.append([
            it.seq,
            it.question,
            it.expected_answer or "",
            it.category or "",
            it.actual_answer or "",
            "是" if it.matched else ("否" if it.matched is not None else ""),
            it.sources_count if it.sources_count is not None else "",
            it.response_time_ms if it.response_time_ms is not None else "",
            it.test_status,
            it.error_msg or "",
            {"correct": "正确", "incorrect": "错误", "pending": "待审核"}.get(it.review_status, "待审核"),
            it.review_reason or "",
        ])

    # 列宽
    widths = [6, 30, 35, 12, 40, 8, 8, 10, 10, 20, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    filename = f"batch_test_{task_id[:8]}_results.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
